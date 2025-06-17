# hashlib: 비밀번호를 해시화하기 위한 표준 라이브러리
import hashlib

# openpyxl: 엑셀 파일을 읽고 쓰기 위한 라이브러리
from openpyxl import Workbook, load_workbook

# os: 파일이 존재하는지 확인하기 위한 모듈
import os

excel_path=".\\authDB\\auth.xlsx"
# 비밀번호를 SHA-256 알고리즘으로 해시하는 함수
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# 엑셀에 새로운 학생의 학번과 해시된 비밀번호를 저장하는 함수
def add_student_to_excel():
    # 입력된 비밀번호를 해시화
    ws,wb = create_excel()
    # 엑셀 파일이 없으면 새로 생성

    ############################################################################################
    auth_dict = make_dictionary()
    for student_id, plain_password in auth_dict.items():
        check_registered_user(student_id,ws)
        # 이미 등록된 학번이 있는지 확인
        hashed_pw = hash_password(plain_password)
        ws.append([student_id, hashed_pw])
        wb.save(excel_path)
        print(f"✅ {student_id} 등록 완료!")

def make_dictionary():
    auth_dict = {"C123000":"1234",
                 "C123400": "0000",
                 "C123450": "hi23423"}
    return auth_dict
    
def create_excel():
    """
    엑셀파일 만드는 함수
    """
    if not os.path.exists(excel_path):
        wb = Workbook()                  # 새 엑셀 워크북 생성
        ws = wb.active                   # 기본 시트 선택
        ws.title = "auth"                # 시트 이름을 "auth"로 설정
        ws.append(["학번", "비밀번호 해시"])  # 첫 줄에 헤더 추가
    else:
        # 엑셀 파일이 있으면 불러오기
        wb = load_workbook(excel_path)
        ws = wb.active
    return ws,wb

def check_registered_user(student_id,ws):
    for row in ws.iter_rows(min_row=2, values_only=True):  # 2번째 줄부터 모든 줄 검사
        if str(row[0]) == str(student_id):  # 학번이 같으면 중복 처리
            print(f"⚠️ 이미 존재하는 학번입니다: {student_id}")
            return  # 함수 종료
add_student_to_excel()